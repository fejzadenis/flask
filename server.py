from flask import Flask, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
PROJECT_FOLDER = 'projects'

@app.route('/')
def home():
    return 'Welcome to your Flask app!'

@app.route('/api/add_project', methods=['POST'])
def add_project():
    try:
        project_data = request.json  # Get JSON data from request

        # Extract project name from the JSON data
        project_name = project_data.get('name')

        # Validate project name (if needed)

        # Create project folder if it doesn't exist
        project_folder = os.path.join(PROJECT_FOLDER, project_name)
        os.makedirs(project_folder, exist_ok=True)

        # Continue processing the request
        filename = f"{project_name}.xlsx"
        filepath = os.path.join(project_folder, filename)

        wb = Workbook()
        ws = wb.active
        ws.append(["Ime projekta", "Datum", "Radni sati", "Radnik", "Opis"])
        wb.save(filepath)

        return jsonify({'message': 'Excel file generated successfully', 'filename': filename}), 201
    except KeyError:
        return jsonify({'error': 'Missing project name'}), 400

@app.route('/api/projects', methods=['GET'])
def get_projects():
    projects = []
    for folder in os.listdir(PROJECT_FOLDER):
        project_folder = os.path.join(PROJECT_FOLDER, folder)
        if os.path.isdir(project_folder):
            excel_files = [os.path.splitext(f)[0] for f in os.listdir(project_folder) if f.endswith('.xlsx') and not f.startswith('~$')]
            projects.extend(excel_files)
    return jsonify({'projects': projects})



@app.route('/api/projects/<string:project_id>/append_data', methods=['POST'])
def append_data_to_project(project_id):
    try:
        data = request.json

        # Iterate over files in the "projects" folder and find the Excel file with the matching name
        excel_file = None  # Initialize the excel_file variable

        for folder_name in os.listdir(PROJECT_FOLDER):
            folder_path = os.path.join(PROJECT_FOLDER, folder_name)
            
            if os.path.isdir(folder_path) and folder_name.startswith(project_id):
                for file_name in os.listdir(folder_path):
                    if file_name.endswith('.xlsx'):
                        excel_file = os.path.join(folder_path, file_name)
                        break  # Assuming you want to break after finding the first Excel file
                if excel_file:
                    break  # Break from the outer loop if Excel file is found
        # Check if an Excel file was found
        if not excel_file:
            return jsonify({'error': f'Excel file for project {project_id} not found'}), 404

        # Construct the filepath of the Excel file
        filepath = excel_file


        # Open the Excel file and append data to the worksheet
        wb = load_workbook(filepath)
        ws = wb.active
        ws.append([
            data.get('name', ''),
            data.get('date', ''),
            data.get('workingHours', ''),
            data.get('employees', ''),
            data.get('description', '')
        ])
        wb.save(filepath)

        return jsonify({'message': 'Data appended successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500




if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')